import xlsxwriter
import os.path
from os import path
import time
import openpyxl

WRITE = 0
READ = 1

class excel:
    def __init__(self):
        print("Create excel file")

    def create_excel_file(self):
        # Create the file
        os_path = os.getcwd()
        print("os_path: ", os_path)
        os_path += '/'
        file_name = os_path + time.strftime("%Y%m%d-%H%M%S")
        file_name += '.xlsx'
        print("Create File Name: ", file_name)
        self.workbook = xlsxwriter.Workbook(file_name)
        self.close_excel_file()

        return file_name

    def set_sheet_name(self, path_file, sheet_name):
        if(path.isfile(path_file)) == False:
            print("FILE Not Exits -> return")
            return

        wb = openpyxl.load_workbook(path_file)

        if sheet_name is not None:
            wb.create_sheet(sheet_name, 0)
            wb.save(path_file)
     
    def write_data(self, path_file, sheet_write, colum_write, row_write, data):
        if(path.isfile(path_file)) == False:
            print("FILE Not Exits -> return")
            return

        xfile = openpyxl.load_workbook(path_file)
        sheet = xfile[sheet_write]

        c1 = sheet.cell(row_write, colum_write)
        c1.value = data
        xfile.save(path_file)

    def close_excel_file(self):
        self.workbook.close()

    def __exit__(self):
        print("Exit Class")

# if __name__== "__main__":

